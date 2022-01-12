#-*-coding:utf8-*-

from pathlib import Path
from typing import Dict, List
from loguru import logger
import yaml
from pyexcel_xlsx import get_data
from openpyxl import load_workbook
from nodes import is_valid_sheet_name
from utils import get_svn_info
from .excel_sheet import ExcelSheet
from .excel_row import ExcelRow

class ExcelFile():
    '''provide a function for parsing excel configuration file into python'''

    def __init__(self) -> None:
        super().__init__()
        self.sheets: Dict[str, ExcelSheet] = {}
        self.yaml_dict: Dict[str, ExcelSheet] = {}
        self.input_excel_path: Path = None

    def save_as_excel(self, excel_path: Path, excel_header: Path):
        '''current current documenets as excel file'''
        workbook = load_workbook(excel_header)
        for sheet_name, sheet in self.sheets.items():
            if sheet_name in workbook:
                logger.debug(f"write to excel {sheet_name} Begin")
                sheet.write_to_excel(workbook[sheet_name])
            else:
                logger.warning(f"write to excel {sheet_name} failed, file not found")
        logger.debug(f"show output folder {excel_path}")
        workbook.save(excel_path)

    def save_as_yaml(self, output_folder: Path):
        '''save current documents as yaml file'''
        for name, sheet in self.sheets.items():
            sheet.save_as_yaml(output_folder/f"{name}.yaml")

    def parse_excel(self, file_path: Path):
        '''parse the excel file given by file path'''
        self.input_excel_path = file_path
        sheets = get_data(str(self.input_excel_path))
        for index, (name, sheet) in enumerate(sheets.items()):
            if not is_valid_sheet_name(name):
                continue
            self.sheets[name] = ExcelSheet.parse_sheet(name, index, sheet)
            self.sheets[name].excel_name = file_path.name

    def add_origin_yaml(self, yaml_list: List[Path]):
        '''add origin yaml list for compare'''
        self.yaml_dict = {}
        for yaml_file in yaml_list:
            self.yaml_dict[yaml_file.absolute().stem] = yaml_file

    def load_yaml(self, yaml_list: List[Path]):
        '''load yaml instead of parse excel'''
        for yaml_file in yaml_list:
            with open(yaml_file, 'r', encoding='utf8') as file:
                self.sheets[yaml_file.absolute().stem] = yaml.unsafe_load(file)
        self.input_excel_path = None

    def compare_yaml_diff(self):
        '''compare yaml from excel and origin yaml, figure out the creator and author'''
        svn_info: Dict = get_svn_info()
        svn_user = svn_info["user"]
        for name, sheet in self.sheets.items():
            yaml_sheet_dict: Dict[str, ExcelRow] = {}
            if name in self.yaml_dict:
                with open(self.yaml_dict[name], 'r', encoding='utf8') as file:
                    # convert row list to id dict, for fast compare later
                    yaml_sheet: ExcelSheet = yaml.unsafe_load(file)
                    for item in yaml_sheet.body:
                        if "id" in item.data:
                            yaml_sheet_dict[item.data["id"]] = item
            # loop sheet to compare row by row, figure out author
            for index, excel_row in enumerate(sheet.body):
                if "id" in excel_row.data:
                    row_id = excel_row.data["id"]
                    if row_id in yaml_sheet_dict:
                        yaml_row = yaml_sheet_dict[row_id]
                        if excel_row == yaml_row:
                            # 一样，那就是没改变，creator和author都是原来的
                            excel_row.creator = yaml_row.creator
                            excel_row.last_author = yaml_row.last_author
                        else:
                            # 不一样，author更新， creator不变
                            excel_row.creator = yaml_row.creator
                            excel_row.last_author = svn_user
                            logger.debug(f"changed: sheet {name} row {index} id {row_id}")
                    else:
                        # 没有，说明是新创建的，creator和author都是当前的
                        excel_row.creator = svn_user
                        excel_row.last_author = svn_user
                        logger.debug(f"Added: sheet {name} row {index} id {row_id}")

    def compare_header(self, sheets):
        '''compare header with input sheet headers, if not equal, replace the header_xls'''
        if len(self.sheets) != len(sheets):
            return False
        # self.sheets.header和sheets比较
        for name, sheet in self.sheets.items():
            if name not in sheets:
                return False
            if len(sheets[name]) != len(sheet.headers):
                return False
            for i, item in enumerate(sheets[name]):
                if item != sheet.headers[i]:
                    return False
        return True

    def replace_header(self, header_xls: Path):
        '''clear the input_excel and use it to replace header_xls if the header is different from parsed data'''
        sheets = get_data(str(header_xls))
        sheet_headers = {}
        for (name, sheet) in sheets.items():
            if not is_valid_sheet_name(name):
                continue
            headers = []
            ExcelSheet.parse_header(sheet[0:5], headers)
            sheet_headers[name] = headers
        if not self.compare_header(sheet_headers):
            logger.warning(f"header changed, save new header from {self.input_excel_path} to {header_xls}")
            if not self.input_excel_path:
                return
            workbook = load_workbook(self.input_excel_path)
            for sheet_name in workbook.sheetnames:
                if workbook[sheet_name].max_row > 5:
                    workbook[sheet_name].delete_rows(6, workbook[sheet_name].max_row - 6 + 1)
            workbook.save(header_xls)
